import copy
import unittest
import xlsxTest
import initData
import json
import pandas
import openpyxl


class TestUnformatted(unittest.TestCase):
    #This test creates a spreadsheet using predetermined data, then compares it to a manually created spreadsheet.
    def test_validate_manual(self):
        xlsxTest.create_manual_spreadsheet()
        automaticTable = pandas.read_excel("plain.xlsx", sheet_name='Sheet1')
        manualTable = pandas.read_excel("manualTable1.xlsx", sheet_name="Sheet1")
        self.assertTrue(automaticTable.equals(manualTable))

    def test_50_random_unformatted(self):
        #Create 50 random datas
        tempJSON = json.loads(initData.create_random_data(50))
        #convert to pandas dataframe

        xlsxTest.create_spreadsheet(tempJSON)
        automaticTable = pandas.read_excel("plain.xlsx", sheet_name='Sheet1')
        autoJSON = {}
        for row in automaticTable.iterrows():
            autoJSON.update({str(row[0]): dict(row[1])})
            del autoJSON[str(row[0])]['Unnamed: 0']

        self.assertTrue(autoJSON == tempJSON)

    def test_200_random_unformatted(self):
        #Create 50 random datas
        tempJSON = json.loads(initData.create_random_data(200))
        #convert to pandas dataframe

        xlsxTest.create_spreadsheet(tempJSON)
        automaticTable = pandas.read_excel("plain.xlsx", sheet_name='Sheet1')
        autoJSON = {}
        for row in automaticTable.iterrows():
            autoJSON.update({str(row[0]): dict(row[1])})
            del autoJSON[str(row[0])]['Unnamed: 0']

        self.assertTrue(autoJSON == tempJSON)

    def test_1000_random_unformatted(self):
        #Create 50 random datas
        tempJSON = json.loads(initData.create_random_data(1000))
        #convert to pandas dataframe

        xlsxTest.create_spreadsheet(tempJSON)
        automaticTable = pandas.read_excel("plain.xlsx", sheet_name='Sheet1')
        autoJSON = {}
        for row in automaticTable.iterrows():
            autoJSON.update({str(row[0]): dict(row[1])})
            del autoJSON[str(row[0])]['Unnamed: 0']

        self.assertTrue(autoJSON == tempJSON)

    def test_10000_random_unformatted(self):
        # Create 50 random datas
        tempJSON = json.loads(initData.create_random_data(10000))
        # convert to pandas dataframe

        xlsxTest.create_spreadsheet(tempJSON)
        automaticTable = pandas.read_excel("plain.xlsx", sheet_name='Sheet1')
        autoJSON = {}
        for row in automaticTable.iterrows():
            autoJSON.update({str(row[0]): dict(row[1])})
            del autoJSON[str(row[0])]['Unnamed: 0']

        self.assertTrue(autoJSON == tempJSON)

    def test_100000_random_unformatted(self):
        # Create 50 random datas
        tempJSON = json.loads(initData.create_random_data(100000))
        # convert to pandas dataframe

        xlsxTest.create_spreadsheet(tempJSON)
        automaticTable = pandas.read_excel("plain.xlsx", sheet_name='Sheet1')
        autoJSON = {}
        for row in automaticTable.iterrows():
            autoJSON.update({str(row[0]): dict(row[1])})
            del autoJSON[str(row[0])]['Unnamed: 0']

        self.assertTrue(autoJSON == tempJSON)

    def test_200000_random_unformatted(self):
        # Create 50 random datas
        tempJSON = json.loads(initData.create_random_data(200000))
        # convert to pandas dataframe

        xlsxTest.create_spreadsheet(tempJSON)
        automaticTable = pandas.read_excel("plain.xlsx", sheet_name='Sheet1')
        autoJSON = {}
        for row in automaticTable.iterrows():
            autoJSON.update({str(row[0]): dict(row[1])})
            del autoJSON[str(row[0])]['Unnamed: 0']

        self.assertTrue(autoJSON == tempJSON)

    def test_300000_random_unformatted(self):
        # Create 50 random datas
        tempJSON = json.loads(initData.create_random_data(300000))
        # convert to pandas dataframe

        xlsxTest.create_spreadsheet(tempJSON)
        automaticTable = pandas.read_excel("plain.xlsx", sheet_name='Sheet1')
        autoJSON = {}
        for row in automaticTable.iterrows():
            autoJSON.update({str(row[0]): dict(row[1])})
            del autoJSON[str(row[0])]['Unnamed: 0']

        self.assertTrue(autoJSON == tempJSON)

    def test_400000_random_unformatted(self):
        # Create 50 random datas
        tempJSON = json.loads(initData.create_random_data(400000))
        # convert to pandas dataframe

        xlsxTest.create_spreadsheet(tempJSON)
        automaticTable = pandas.read_excel("plain.xlsx", sheet_name='Sheet1')
        autoJSON = {}
        for row in automaticTable.iterrows():
            autoJSON.update({str(row[0]): dict(row[1])})
            del autoJSON[str(row[0])]['Unnamed: 0']

        self.assertTrue(autoJSON == tempJSON)

    def test_500000_random_unformatted(self):
        # Create 50 random datas
        tempJSON = json.loads(initData.create_random_data(500000))
        # convert to pandas dataframe

        xlsxTest.create_spreadsheet(tempJSON)
        automaticTable = pandas.read_excel("plain.xlsx", sheet_name='Sheet1')
        autoJSON = {}
        for row in automaticTable.iterrows():
            autoJSON.update({str(row[0]): dict(row[1])})
            del autoJSON[str(row[0])]['Unnamed: 0']

        self.assertTrue(autoJSON == tempJSON)

    def test_1000000_random_unformatted(self):
        # Create 50 random datas
        tempJSON = json.loads(initData.create_random_data(1000000))
        # convert to pandas dataframe

        xlsxTest.create_spreadsheet(tempJSON)
        automaticTable = pandas.read_excel("plain.xlsx", sheet_name='Sheet1')
        autoJSON = {}
        for row in automaticTable.iterrows():
            autoJSON.update({str(row[0]): dict(row[1])})
            del autoJSON[str(row[0])]['Unnamed: 0']

        self.assertTrue(autoJSON == tempJSON)


class TestSymbolic(unittest.TestCase):
    def test_validate_symbolic(self):
        xlsxTest.create_sym_spreadsheet()
        automaticTable = pandas.read_excel("symbolic.xlsx", sheet_name='Sheet1')
        manualTable = pandas.read_excel("manualTable3.xlsx", sheet_name="Sheet1")
        self.assertTrue(automaticTable.equals(manualTable))

    def test_50_random_symbolic(self):
        tempJSON = json.loads(initData.create_random_data(50))
        originalJSON = copy.deepcopy(tempJSON)
        xlsxTest.create_auto_sym_spreadsheet(tempJSON)
        pyxlAutoTable = openpyxl.load_workbook("symbolic.xlsx")
        pyxlAutoTable = pyxlAutoTable["Sheet1"]
        automaticTable = pandas.read_excel("symbolic.xlsx", sheet_name='Sheet1')

        autoJSON = {}
        #Iterate through table loaded in openpyxl to get comments
        DateTime = {}
        for row in pyxlAutoTable.iter_rows():
            if row[3].comment:
                tempDate = row[3].comment.text.split(" ")[0]
                tempTime = row[3].comment.text.split(" ")[1]
                DateTime.update({row[3].row - 2: {"Date": tempDate, "Time": tempTime}})

        for row in automaticTable.iterrows():
            autoJSON.update({str(row[0]): dict(row[1])})
            autoJSON[str(row[0])]['Date'] = DateTime[int(row[0])]['Date']
            autoJSON[str(row[0])]['Time'] = DateTime[int(row[0])]['Time']
            del autoJSON[str(row[0])]['Timestamp']
            del autoJSON[str(row[0])]['Unnamed: 0']

        self.assertTrue(autoJSON == originalJSON)

    def test_200_random_symbolic(self):
        tempJSON = json.loads(initData.create_random_data(200))
        originalJSON = copy.deepcopy(tempJSON)
        xlsxTest.create_auto_sym_spreadsheet(tempJSON)
        pyxlAutoTable = openpyxl.load_workbook("symbolic.xlsx")
        pyxlAutoTable = pyxlAutoTable["Sheet1"]
        automaticTable = pandas.read_excel("symbolic.xlsx", sheet_name='Sheet1')

        autoJSON = {}
        # Iterate through table loaded in openpyxl to get comments
        DateTime = {}
        for row in pyxlAutoTable.iter_rows():
            if row[3].comment:
                tempDate = row[3].comment.text.split(" ")[0]
                tempTime = row[3].comment.text.split(" ")[1]
                DateTime.update({row[3].row - 2: {"Date": tempDate, "Time": tempTime}})

        for row in automaticTable.iterrows():
            autoJSON.update({str(row[0]): dict(row[1])})
            autoJSON[str(row[0])]['Date'] = DateTime[int(row[0])]['Date']
            autoJSON[str(row[0])]['Time'] = DateTime[int(row[0])]['Time']
            del autoJSON[str(row[0])]['Timestamp']
            del autoJSON[str(row[0])]['Unnamed: 0']

        self.assertTrue(autoJSON == originalJSON)

    def test_1000_random_symbolic(self):
        tempJSON = json.loads(initData.create_random_data(1000))
        originalJSON = copy.deepcopy(tempJSON)
        xlsxTest.create_auto_sym_spreadsheet(tempJSON)
        pyxlAutoTable = openpyxl.load_workbook("symbolic.xlsx")
        pyxlAutoTable = pyxlAutoTable["Sheet1"]
        automaticTable = pandas.read_excel("symbolic.xlsx", sheet_name='Sheet1')

        autoJSON = {}
        # Iterate through table loaded in openpyxl to get comments
        DateTime = {}
        for row in pyxlAutoTable.iter_rows():
            if row[3].comment:
                tempDate = row[3].comment.text.split(" ")[0]
                tempTime = row[3].comment.text.split(" ")[1]
                DateTime.update({row[3].row - 2: {"Date": tempDate, "Time": tempTime}})

        for row in automaticTable.iterrows():
            autoJSON.update({str(row[0]): dict(row[1])})
            autoJSON[str(row[0])]['Date'] = DateTime[int(row[0])]['Date']
            autoJSON[str(row[0])]['Time'] = DateTime[int(row[0])]['Time']
            del autoJSON[str(row[0])]['Timestamp']
            del autoJSON[str(row[0])]['Unnamed: 0']

        self.assertTrue(autoJSON == originalJSON)

    def test_10000_random_symbolic(self):
        tempJSON = json.loads(initData.create_random_data(10000))
        originalJSON = copy.deepcopy(tempJSON)
        xlsxTest.create_auto_sym_spreadsheet(tempJSON)
        pyxlAutoTable = openpyxl.load_workbook("symbolic.xlsx")
        pyxlAutoTable = pyxlAutoTable["Sheet1"]
        automaticTable = pandas.read_excel("symbolic.xlsx", sheet_name='Sheet1')

        autoJSON = {}
        # Iterate through table loaded in openpyxl to get comments
        DateTime = {}
        for row in pyxlAutoTable.iter_rows():
            if row[3].comment:
                tempDate = row[3].comment.text.split(" ")[0]
                tempTime = row[3].comment.text.split(" ")[1]
                DateTime.update({row[3].row - 2: {"Date": tempDate, "Time": tempTime}})

        for row in automaticTable.iterrows():
            autoJSON.update({str(row[0]): dict(row[1])})
            autoJSON[str(row[0])]['Date'] = DateTime[int(row[0])]['Date']
            autoJSON[str(row[0])]['Time'] = DateTime[int(row[0])]['Time']
            del autoJSON[str(row[0])]['Timestamp']
            del autoJSON[str(row[0])]['Unnamed: 0']

        self.assertTrue(autoJSON == originalJSON)

    def test_100000_random_symbolic(self):
        tempJSON = json.loads(initData.create_random_data(100000))
        originalJSON = copy.deepcopy(tempJSON)
        xlsxTest.create_auto_sym_spreadsheet(tempJSON)
        print("Validating Spreadsheet")
        pyxlAutoTable = openpyxl.load_workbook("symbolic.xlsx")
        pyxlAutoTable = pyxlAutoTable["Sheet1"]
        automaticTable = pandas.read_excel("symbolic.xlsx", sheet_name='Sheet1')

        autoJSON = {}
        # Iterate through table loaded in openpyxl to get comments
        DateTime = {}
        for row in pyxlAutoTable.iter_rows():
            if row[3].comment:
                tempDate = row[3].comment.text.split(" ")[0]
                tempTime = row[3].comment.text.split(" ")[1]
                DateTime.update({row[3].row - 2: {"Date": tempDate, "Time": tempTime}})

        for row in automaticTable.iterrows():
            autoJSON.update({str(row[0]): dict(row[1])})
            autoJSON[str(row[0])]['Date'] = DateTime[int(row[0])]['Date']
            autoJSON[str(row[0])]['Time'] = DateTime[int(row[0])]['Time']
            del autoJSON[str(row[0])]['Timestamp']
            del autoJSON[str(row[0])]['Unnamed: 0']

        self.assertTrue(autoJSON == originalJSON)

    def test_200000_random_symbolic(self):
        tempJSON = json.loads(initData.create_random_data(200000))
        originalJSON = copy.deepcopy(tempJSON)
        xlsxTest.create_auto_sym_spreadsheet(tempJSON)
        print("Validating Spreadsheet")
        pyxlAutoTable = openpyxl.load_workbook("symbolic.xlsx")
        pyxlAutoTable = pyxlAutoTable["Sheet1"]
        automaticTable = pandas.read_excel("symbolic.xlsx", sheet_name='Sheet1')

        autoJSON = {}
        # Iterate through table loaded in openpyxl to get comments
        DateTime = {}
        for row in pyxlAutoTable.iter_rows():
            if row[3].comment:
                tempDate = row[3].comment.text.split(" ")[0]
                tempTime = row[3].comment.text.split(" ")[1]
                DateTime.update({row[3].row - 2: {"Date": tempDate, "Time": tempTime}})

        for row in automaticTable.iterrows():
            autoJSON.update({str(row[0]): dict(row[1])})
            autoJSON[str(row[0])]['Date'] = DateTime[int(row[0])]['Date']
            autoJSON[str(row[0])]['Time'] = DateTime[int(row[0])]['Time']
            del autoJSON[str(row[0])]['Timestamp']
            del autoJSON[str(row[0])]['Unnamed: 0']

        self.assertTrue(autoJSON == originalJSON)

    def test_300000_random_symbolic(self):
        tempJSON = json.loads(initData.create_random_data(300000))
        originalJSON = copy.deepcopy(tempJSON)
        xlsxTest.create_auto_sym_spreadsheet(tempJSON)
        print("Validating Spreadsheet")
        pyxlAutoTable = openpyxl.load_workbook("symbolic.xlsx")
        pyxlAutoTable = pyxlAutoTable["Sheet1"]
        automaticTable = pandas.read_excel("symbolic.xlsx", sheet_name='Sheet1')

        autoJSON = {}
        # Iterate through table loaded in openpyxl to get comments
        DateTime = {}
        for row in pyxlAutoTable.iter_rows():
            if row[3].comment:
                tempDate = row[3].comment.text.split(" ")[0]
                tempTime = row[3].comment.text.split(" ")[1]
                DateTime.update({row[3].row - 2: {"Date": tempDate, "Time": tempTime}})

        for row in automaticTable.iterrows():
            autoJSON.update({str(row[0]): dict(row[1])})
            autoJSON[str(row[0])]['Date'] = DateTime[int(row[0])]['Date']
            autoJSON[str(row[0])]['Time'] = DateTime[int(row[0])]['Time']
            del autoJSON[str(row[0])]['Timestamp']
            del autoJSON[str(row[0])]['Unnamed: 0']

        self.assertTrue(autoJSON == originalJSON)

    def test_400000_random_symbolic(self):
        tempJSON = json.loads(initData.create_random_data(400000))
        originalJSON = copy.deepcopy(tempJSON)
        xlsxTest.create_auto_sym_spreadsheet(tempJSON)
        print("Validating Spreadsheet")
        pyxlAutoTable = openpyxl.load_workbook("symbolic.xlsx")
        pyxlAutoTable = pyxlAutoTable["Sheet1"]
        automaticTable = pandas.read_excel("symbolic.xlsx", sheet_name='Sheet1')

        autoJSON = {}
        # Iterate through table loaded in openpyxl to get comments
        DateTime = {}
        for row in pyxlAutoTable.iter_rows():
            if row[3].comment:
                tempDate = row[3].comment.text.split(" ")[0]
                tempTime = row[3].comment.text.split(" ")[1]
                DateTime.update({row[3].row - 2: {"Date": tempDate, "Time": tempTime}})

        for row in automaticTable.iterrows():
            autoJSON.update({str(row[0]): dict(row[1])})
            autoJSON[str(row[0])]['Date'] = DateTime[int(row[0])]['Date']
            autoJSON[str(row[0])]['Time'] = DateTime[int(row[0])]['Time']
            del autoJSON[str(row[0])]['Timestamp']
            del autoJSON[str(row[0])]['Unnamed: 0']

        self.assertTrue(autoJSON == originalJSON)

    def test_500000_random_symbolic(self):
        tempJSON = json.loads(initData.create_random_data(500000))
        originalJSON = copy.deepcopy(tempJSON)
        xlsxTest.create_auto_sym_spreadsheet(tempJSON)
        print("Validating Spreadsheet")
        pyxlAutoTable = openpyxl.load_workbook("symbolic.xlsx")
        pyxlAutoTable = pyxlAutoTable["Sheet1"]
        automaticTable = pandas.read_excel("symbolic.xlsx", sheet_name='Sheet1')

        autoJSON = {}
        # Iterate through table loaded in openpyxl to get comments
        DateTime = {}
        for row in pyxlAutoTable.iter_rows():
            if row[3].comment:
                tempDate = row[3].comment.text.split(" ")[0]
                tempTime = row[3].comment.text.split(" ")[1]
                DateTime.update({row[3].row - 2: {"Date": tempDate, "Time": tempTime}})

        for row in automaticTable.iterrows():
            autoJSON.update({str(row[0]): dict(row[1])})
            autoJSON[str(row[0])]['Date'] = DateTime[int(row[0])]['Date']
            autoJSON[str(row[0])]['Time'] = DateTime[int(row[0])]['Time']
            del autoJSON[str(row[0])]['Timestamp']
            del autoJSON[str(row[0])]['Unnamed: 0']

        self.assertTrue(autoJSON == originalJSON)

    def test_1000000_random_symbolic(self):
        tempJSON = json.loads(initData.create_random_data(1000000))
        originalJSON = copy.deepcopy(tempJSON)
        xlsxTest.create_auto_sym_spreadsheet(tempJSON)
        print("Validating Spreadsheet")
        pyxlAutoTable = openpyxl.load_workbook("symbolic.xlsx")
        pyxlAutoTable = pyxlAutoTable["Sheet1"]
        automaticTable = pandas.read_excel("symbolic.xlsx", sheet_name='Sheet1')

        autoJSON = {}
        # Iterate through table loaded in openpyxl to get comments
        DateTime = {}
        for row in pyxlAutoTable.iter_rows():
            if row[3].comment:
                tempDate = row[3].comment.text.split(" ")[0]
                tempTime = row[3].comment.text.split(" ")[1]
                DateTime.update({row[3].row - 2: {"Date": tempDate, "Time": tempTime}})

        for row in automaticTable.iterrows():
            autoJSON.update({str(row[0]): dict(row[1])})
            autoJSON[str(row[0])]['Date'] = DateTime[int(row[0])]['Date']
            autoJSON[str(row[0])]['Time'] = DateTime[int(row[0])]['Time']
            del autoJSON[str(row[0])]['Timestamp']
            del autoJSON[str(row[0])]['Unnamed: 0']

        self.assertTrue(autoJSON == originalJSON)


if __name__ == '__main__':
    unittest.main()
